import yfinance as yf
import pandas as pd
import os
from openpyxl import load_workbook
import streamlit as st

# Path to the stocks.xlsx file
STOCKS_FILE_PATH = 'stocks.xlsx'  # Change this to the correct path if needed

# Function to fetch data for a given stock ticker
def get_financial_data(ticker):
    stock = yf.Ticker(ticker)
    result = {'Ticker': ticker}
    
    try:
        income_statement = stock.financials
        balance_sheet = stock.balance_sheet
        cash_flow = stock.cashflow
        dividends = stock.dividends
    except Exception as e:
        st.error(f"Error fetching financial data for {ticker}: {e}")
        return None

    try:
        historical_data = stock.history(period="1d")
        latest_close_price = historical_data['Close'].iloc[-1]
    except Exception as e:
        latest_close_price = "N/A"

    result['Net Income'] = income_statement.loc['Net Income'] if 'Net Income' in income_statement.index else "N/A"
    result['Operating Income'] = income_statement.loc['Operating Income'] if 'Operating Income' in income_statement.index else \
                                 income_statement.loc['EBIT'] if 'EBIT' in income_statement.index else "N/A"
    
    try:
        eps = income_statement.loc['Earnings Before Interest and Taxes'] / stock.info['sharesOutstanding']
    except KeyError:
        eps = "N/A"
    result['EPS'] = eps
    
    result['Revenue Growth'] = income_statement.loc['Total Revenue'].pct_change().iloc[-1] if 'Total Revenue' in income_statement.index else "N/A"
    
    result['Retained Earnings'] = balance_sheet.loc['Retained Earnings'] if 'Retained Earnings' in balance_sheet.index else "N/A"
    result['Cash Reserves'] = balance_sheet.loc['Cash'] if 'Cash' in balance_sheet.index else "N/A"
    
    try:
        result['Debt-to-Equity Ratio'] = balance_sheet.loc['Total Debt'] / balance_sheet.loc['Stockholders Equity'] if 'Total Debt' in balance_sheet.index and 'Stockholders Equity' in balance_sheet.index else "N/A"
    except KeyError:
        result['Debt-to-Equity Ratio'] = "N/A"
    
    result['Working Capital'] = balance_sheet.loc['Total Assets'] - balance_sheet.loc['Total Liabilities Net Minority Interest'] if 'Total Assets' in balance_sheet.index and 'Total Liabilities Net Minority Interest' in balance_sheet.index else "N/A"
    
    result['Dividend Payout Ratio'] = stock.info.get('dividendYield', "N/A")
    result['Dividend Yield'] = result['Dividend Payout Ratio']
    
    result['Free Cash Flow'] = cash_flow.loc['Free Cash Flow'] if 'Free Cash Flow' in cash_flow.index else "N/A"
    
    if not dividends.empty:
        result['Dividend Growth Rate'] = dividends.pct_change().mean()
    else:
        result['Dividend Growth Rate'] = "N/A"
    
    result['Latest Close Price'] = latest_close_price
    result['Dividend Percentage'] = "N/A"
    
    if not dividends.empty:
        predicted_dividend_amount = dividends.iloc[-1]
        if latest_close_price != "N/A":
            dividend_percentage = (predicted_dividend_amount / latest_close_price) * 100
            result['Dividend Percentage'] = dividend_percentage
        
        past_dividends = dividends.tail(10)
        result['Past Dividends'] = past_dividends.tolist()
        
        date_diffs = past_dividends.index.to_series().diff().dropna()
        if not date_diffs.empty:
            avg_diff = date_diffs.mean()
            last_dividend_date = past_dividends.index[-1]
            next_dividend_date = last_dividend_date + avg_diff
            result['Next Dividend Date'] = str(next_dividend_date)
        else:
            result['Next Dividend Date'] = 'N/A'

        result['Predicted Dividend Amount'] = predicted_dividend_amount
    else:
        result['Next Dividend Date'] = 'N/A'
        result['Predicted Dividend Amount'] = 'N/A'
        result['Dividend Percentage'] = "N/A"

    return result

# Function to save results to an Excel file
def save_to_excel(results, filename="dividend_predictions.xlsx"):
    try:
        results_df = pd.DataFrame(results)
        if os.path.exists(filename):
            book = load_workbook(filename)
            writer = pd.ExcelWriter(filename, engine='openpyxl')
            writer.book = book
            results_df.to_excel(writer, index=False, header=False, startrow=book.active.max_row)
            writer.save()
        else:
            results_df.to_excel(filename, index=False)
        st.success(f"Results saved to {filename}")
    except Exception as e:
        st.error(f"Error saving to Excel: {e}")

# Streamlit App
st.set_page_config(page_title="Stock Dividend Predictionss", layout="wide")

# Display Header Logo
st.markdown("""
    <style>
        .header-logo {
            display: block;
            margin-left: auto;
            margin-right: auto;
            width: 25%;
        }
        /* Hide GitHub icons and fork button */
        .css-1v0mbdj { 
            display: none !important;
        }
        .css-1b22hs3 {
            display: none !important;
        }
        /* Hide Streamlit footer elements */
        footer { 
            display: none !important; 
        }
        /* Hide the GitHub repository button */
        .css-1r6ntm8 { 
            display: none !important;
        }
    </style>
    <img class="header-logo" src="https://pystatiq.com/images/pystatIQ_logo.png" alt="Header Logo">
""", unsafe_allow_html=True)

st.title('Stock Dividend Prediction and Financial Analysis')

# Read the stock symbols from the local stocks.xlsx file
if os.path.exists(STOCKS_FILE_PATH):
    symbols_df = pd.read_excel(STOCKS_FILE_PATH)

    # Check if the 'Symbol' column exists
    if 'Symbol' not in symbols_df.columns:
        st.error("The file must contain a 'Symbol' column with stock tickers.")
    else:
        # Let the user select stocks from the file
        stock_options = symbols_df['Symbol'].tolist()
        selected_stocks = st.multiselect("Select Stock Symbols", stock_options)

        # Button to start the data fetching process
        if st.button('Fetch Financial Data') and selected_stocks:
            all_results = []
            for ticker in selected_stocks:  # Use selected stocks
                st.write(f"Processing {ticker}...")
                result = get_financial_data(ticker)
                if result is not None:
                    all_results.append(result)
            
            if all_results:
                st.subheader("Financial Data Results")
                results_df = pd.DataFrame(all_results)
                st.dataframe(results_df)
                
                # Button to save the results to Excel
                if st.button('Save Results to Excel'):
                    save_to_excel(all_results)

else:
    st.error(f"{STOCKS_FILE_PATH} not found. Please ensure the file exists.")

# Display Footer Logo
# Content before the footer logo
st.markdown("""
    <div style="text-align: center; font-size: 14px; margin-top: 30px;">
        <p><strong>App Code:</strong> Stock-Dividend-Prediction-Jan-2025</p>
        <p>To get access to the stocks file to upload, please Email us at <a href="mailto:support@pystatiq.com">support@pystatiq.com</a>.</p>
        <p>Don't forget to add the Application code.</p>
        <p><strong>README:</strong> <a href="https://pystatiq-lab.gitbook.io/docs/python-apps/stock-dividend-predictions" target="_blank">https://pystatiq-lab.gitbook.io/docs/python-apps/stock-dividend-predictions</a></p>
    </div>
""", unsafe_allow_html=True)

# Display Footer Logo
st.markdown(f"""
    <style>
        .footer-logo {{
            display: block;
            margin-left: auto;
            margin-right: auto;
            width: 90px;
            padding-top: 30px;
        }}
    </style>
    <img class="footer-logo" src="https://predictram.com/images/logo.png" alt="Footer Logo">
""", unsafe_allow_html=True)
